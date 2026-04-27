"""Generate fb-groups-to-post.xlsx, two sheets:

Sheet 1, "FB Groups (Hosts)": Miami-Dade and Broward groups for the
marketplace pitch (people who own/operate space and could list it).

Sheet 2, "SaaS Studio Software (National)": national groups of people
who already have their own clients and could use Align's Studio
Software as their booking + scheduling system.
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

HOST_URL = "alignworkspaces.com/for-hosts"
SAAS_URL = "alignworkspaces.com/for-studios"

# Image filenames (in documents/marketing/fb-post-images/)
IMG_HOST_MIAMI = "host-miami.jpg"
IMG_HOST_BROWARD = "host-broward.jpg"
IMG_HOST_INVESTOR = "host-investor.jpg"
IMG_SAAS_SALON = "saas-salon.jpg"
IMG_SAAS_MEDSPA = "saas-medspa.jpg"
IMG_SAAS_FITNESS = "saas-fitness.jpg"
IMG_SAAS_COACH = "saas-coach.jpg"
IMG_SAAS_THERAPY = "saas-therapy.jpg"
IMG_SAAS_PHOTO = "saas-photo.jpg"
IMG_SAAS_RECORDING = "saas-recording.jpg"

# === Sheet 1: Miami-Dade + Broward HOST groups ===
# (Name, FB URL, County, Audience, Pitch, Notes, Link, Image, Post)
HOST_GROUPS = [
    ("Small Business Owners of Miami & Fort Lauderdale",
     "https://www.facebook.com/groups/1157985140883947/",
     "Miami-Dade + Broward",
     "Miami + Ft Lauderdale small business owners",
     "Marketplace: list any commercial space for empty hours",
     "Broad audience, many own or lease commercial space.",
     HOST_URL, IMG_HOST_MIAMI,
     "Hey Miami + Broward business owners, my wife and I just launched alignworkspaces.com/for-hosts and wanted to share it with this group.\n\n"
     "If you've got commercial space (office, meeting room, treatment room, studio) that sits empty for parts of the week, you can list those hours and earn from them. No upfront cost, you only pay anything when it actually books.\n\n"
     "We built it because we noticed a lot of small businesses around here have great spaces sitting half-used. Mods, please take this down if it isn't allowed."),

    ("South Florida Small Business Network",
     "https://www.facebook.com/groups/SouthFloridaSmallBizNetwork/",
     "Miami-Dade + Broward",
     "South Florida small businesses",
     "Marketplace: monetize empty commercial hours",
     "Networking-focused, soft post.",
     HOST_URL, IMG_HOST_MIAMI,
     "Hey South FL business community, my wife and I just launched alignworkspaces.com/for-hosts and wanted to share it with this group.\n\n"
     "Idea is simple: if you've got a professional space sitting empty part of the week, list those hours and earn from them. Free to list, no monthly fee.\n\n"
     "Wanted to put it in front of this network first. Open to any feedback, and mods, please remove if this isn't allowed."),

    ("Miami Business Network",
     "https://www.facebook.com/groups/520790141319929/",
     "Miami-Dade",
     "Miami business owners",
     "Marketplace + Studio Software",
     "Networking group, soft sell.",
     HOST_URL, IMG_HOST_MIAMI,
     "Hey everyone, my wife and I just launched alignworkspaces.com/for-hosts and wanted to share with this group.\n\n"
     "Two things you can do: list any commercial space you own and earn from the empty hours, or if you run a service business with your own clients (salon, therapy practice, photo studio, wellness, etc.), use our software to handle bookings, payments, and your own branded booking page.\n\n"
     "Curious what people in this group think. Mods, take this down if it doesn't belong."),

    ("Broward County Business and Networking Group",
     "https://www.facebook.com/groups/browardbusiness/",
     "Broward",
     "Broward business owners",
     "Marketplace + Studio Software",
     "Primary Broward business group.",
     HOST_URL, IMG_HOST_BROWARD,
     "Hey Broward, my wife and I just launched alignworkspaces.com/for-hosts and wanted to share it here.\n\n"
     "If you've got commercial space that's not in use 24/7 (offices, meeting rooms, studios, treatment rooms), you can list the empty hours and earn from them. If you run a service practice with your own clients, you can also use our software to handle bookings and payments under your own brand.\n\n"
     "Wanted to put it in front of this group for feedback. Mods, please remove if it isn't a fit."),

    ("MIAMI WOMEN ENTREPRENEURS",
     "https://www.facebook.com/groups/MiamiWomenEntrepreneurs/",
     "Miami-Dade",
     "Miami women business owners (many in beauty, wellness, coaching)",
     "Marketplace + Studio Software",
     "Strong host-fit demographic.",
     HOST_URL, IMG_HOST_MIAMI,
     "Hey ladies, my wife and I just launched alignworkspaces.com/for-hosts and wanted to share it with this group.\n\n"
     "If you own a studio, salon, treatment room, or any kind of professional space and it sits empty for parts of the week, you can list those hours and earn from them. If you've got your own clients and just need better software to run bookings (calendar, payments, branded booking page), we have that side too.\n\n"
     "Wanted to put it in front of this community first since so many of you run beautiful local spaces. Mods, please take this down if it isn't allowed."),

    ("Miami Real Estate Investor",
     "https://www.facebook.com/groups/MiamiRealEstateInvestor/",
     "Miami-Dade",
     "Miami real estate investors",
     "Marketplace: monetize vacant commercial inventory",
     "Investor angle.",
     HOST_URL, IMG_HOST_INVESTOR,
     "Hey Miami investors, my wife and I just launched alignworkspaces.com/for-hosts and wanted to share it with this group.\n\n"
     "If you've got commercial property sitting between tenants (or partially occupied), you can list those vacant hours and rent them out by the hour. No setup cost, you only pay anything when bookings come in. Turn the dead months into actual cashflow.\n\n"
     "Wanted to put it in front of this group for feedback. Mods, take this down if it isn't allowed."),

    ("Miami Real Estate Investors",
     "https://www.facebook.com/groups/1548432032097696/",
     "Miami-Dade",
     "Miami real estate investors",
     "Marketplace: monetize vacant commercial inventory",
     "Same angle as above.",
     HOST_URL, IMG_HOST_INVESTOR,
     "Hey everyone, my wife and I just launched alignworkspaces.com/for-hosts and wanted to drop it here.\n\n"
     "For investors with vacant commercial inventory: list those empty hours, earn from the downtime. Free to list, you only pay anything when it actually books.\n\n"
     "Open to feedback from anyone in this group who's dealt with vacant property. Mods, please remove if not allowed."),

    ("South Florida Real Estate Investors",
     "https://www.facebook.com/groups/117204431658948/",
     "Miami-Dade + Broward",
     "South Florida real estate investors",
     "Marketplace: monetize vacant commercial inventory",
     "Larger SoFL audience.",
     HOST_URL, IMG_HOST_INVESTOR,
     "Hey South FL investors, my wife and I just launched alignworkspaces.com/for-hosts and wanted to share it with this group.\n\n"
     "If you've got commercial property sitting empty between tenants, you can list those hours and rent them out by the hour. No upfront cost, no monthly fee. You only pay anything when bookings actually come in.\n\n"
     "Would love feedback from this community. Mods, please remove if it isn't a fit."),

    ("Florida Real Estate Investors / Cash Buyer",
     "https://www.facebook.com/groups/939189196804053/",
     "Statewide FL (incl. Miami-Dade + Broward)",
     "Florida real estate investors",
     "Marketplace: monetize vacant commercial inventory",
     "Statewide. Most active in SoFL.",
     HOST_URL, IMG_HOST_INVESTOR,
     "Hey FL investors, my wife and I just launched alignworkspaces.com/for-hosts and wanted to share with this group.\n\n"
     "If you've got commercial inventory in Miami or Broward sitting between tenants, you can now list those empty hours by the hour. Free to list, only pay when bookings come in.\n\n"
     "Currently focused on Miami-Dade and Broward, would love feedback from anyone in the state who's dealt with vacant commercial. Mods, take this down if it isn't allowed."),

    ("Broward County Real Estate",
     "https://www.facebook.com/groups/100176953850358/",
     "Broward",
     "Broward real estate community (investor-focused)",
     "Marketplace: monetize vacant commercial inventory",
     "Mixed audience reframed for investor angle.",
     HOST_URL, IMG_HOST_INVESTOR,
     "Hey Broward, my wife and I just launched alignworkspaces.com/for-hosts and wanted to share with this group.\n\n"
     "If you own commercial property here that's vacant between tenants or sitting unused part of the week, you can list those hours on alignworkspaces.com and rent them out by the hour. No setup cost, no monthly fee, just pay when bookings come in.\n\n"
     "Wanted to put it in front of this community for feedback. Mods, take this down if it isn't a good fit."),

    ("Wynwood Business Networking",
     "https://www.facebook.com/groups/729772892942746/",
     "Miami-Dade (Wynwood)",
     "Wynwood neighborhood business owners",
     "Marketplace: list space, especially creative/photo/event-friendly Wynwood spaces",
     "Hyperlocal, strong creative + small biz mix.",
     HOST_URL, IMG_HOST_MIAMI,
     "Hey Wynwood, my wife and I just launched alignworkspaces.com/for-hosts and wanted to share it here.\n\n"
     "If you've got commercial space in Wynwood that sits empty for parts of the week, you can list those hours and earn from them. No upfront cost. Plenty of creatives, photographers, and small businesses would love to book Wynwood spaces by the hour.\n\n"
     "Wanted to put it in front of this group first. Mods, please remove if it isn't allowed."),

    ("We Love Pembroke Pines",
     "https://www.facebook.com/groups/WeLovePembrokePines/",
     "Broward (Pembroke Pines)",
     "Pembroke Pines locals + business owners",
     "Marketplace: list any commercial space sitting empty",
     "Community group, includes many local biz owners.",
     HOST_URL, IMG_HOST_BROWARD,
     "Hey Pembroke Pines, my wife and I just launched alignworkspaces.com/for-hosts and wanted to share it with the community.\n\n"
     "If you own commercial space (office, treatment room, studio, meeting room) and it sits empty part of the week, you can list those hours and earn from them. Free to list, only pay anything if it actually books.\n\n"
     "Wanted to put it in front of locals first. Mods, please remove if this isn't allowed."),

    ("Florida Small Business Network",
     "https://www.facebook.com/groups/floridasmallbusinessnetwork/",
     "Statewide FL",
     "Florida small business owners",
     "Marketplace: list commercial space in Miami/Broward",
     "Statewide but plenty of SoFL members.",
     HOST_URL, IMG_HOST_MIAMI,
     "Hey FL small business owners, my wife and I just launched alignworkspaces.com/for-hosts and wanted to share it with this group.\n\n"
     "We're focused on Miami-Dade and Broward right now. If you own commercial space in either county that sits unused part of the week, you can list those hours on alignworkspaces.com and earn from them. No upfront cost.\n\n"
     "Open to feedback from anyone in the network. Mods, take this down if not allowed."),

    ("Business Owners in Florida",
     "https://www.facebook.com/groups/businessownersinflorida/",
     "Statewide FL",
     "Florida business owners",
     "Marketplace: list commercial space in Miami/Broward",
     "Statewide.",
     HOST_URL, IMG_HOST_MIAMI,
     "Hey Florida business owners, my wife and I just launched alignworkspaces.com/for-hosts and wanted to put it in front of this group.\n\n"
     "If you've got commercial space in the Miami or Broward area that's not in full use, you can list it by the hour and earn from the empty time. Free to list, no monthly fee.\n\n"
     "Curious what FL owners think. Mods, please remove if this isn't a good fit."),

    ("Latino Small Business Owners (LSBO)",
     "https://www.facebook.com/groups/LatinoSBO/",
     "National (heavy Miami presence)",
     "Latino small business owners",
     "Marketplace: list space in Miami/Broward",
     "National but Miami is largest demographic.",
     HOST_URL, IMG_HOST_MIAMI,
     "Hey LSBO, my wife and I just launched alignworkspaces.com/for-hosts and wanted to share it with this group.\n\n"
     "We're starting in Miami-Dade and Broward, where so much of this community runs businesses. If you own commercial space and it's empty part of the week, you can list those hours and earn from them. No setup cost.\n\n"
     "Wanted to put it in front of this network first. Mods, take this down if it isn't allowed."),

    ("Miami & Florida Black Owned Businesses",
     "https://www.facebook.com/groups/1628161307430309/",
     "Statewide FL (Miami-focused)",
     "Black-owned Miami + FL businesses",
     "Marketplace: list commercial space",
     "Active Miami/FL group.",
     HOST_URL, IMG_HOST_MIAMI,
     "Hey everyone, my wife and I just launched alignworkspaces.com/for-hosts and wanted to share it with this group.\n\n"
     "If you own commercial space in Miami or Broward (office, salon, studio, meeting room, treatment room) and have empty hours, you can list those and earn from the downtime. Free to list, no monthly.\n\n"
     "Wanted to put it in front of this community first. Mods, please remove if this isn't allowed."),

    ("Women Entrepreneurs of South Florida",
     "https://www.facebook.com/groups/769289819758497/",
     "Miami-Dade + Broward",
     "South Florida women business owners",
     "Marketplace + Studio Software",
     "SoFL women in business, many in beauty/wellness/coaching.",
     HOST_URL, IMG_HOST_MIAMI,
     "Hey ladies, my wife and I just launched alignworkspaces.com/for-hosts and wanted to share it with this group.\n\n"
     "If you run a studio, salon, treatment room, or any kind of professional space in Miami or Broward and it's not full all week, you can list those empty hours and earn from them. If you run a service business with your own clients, our software side handles bookings, payments, and a branded booking page.\n\n"
     "Wanted to put it in front of this community first. Mods, take this down if it isn't allowed."),

    # === New additions (round 3) ===
    ("Miami Business Networking Group",
     "https://www.facebook.com/groups/JOINMBNG/",
     "Miami-Dade",
     "Miami business networking community",
     "Marketplace + Studio Software",
     "Active Miami networking group.",
     HOST_URL, IMG_HOST_MIAMI,
     "Hey Miami, my wife and I just launched alignworkspaces.com/for-hosts and wanted to share it with this group.\n\n"
     "If you own commercial space in Miami and it sits empty for parts of the week, you can list those hours and earn from them. No upfront cost, no monthly fee.\n\n"
     "Wanted to put it in front of this group for feedback. Mods, please remove if it isn't allowed."),

    ("South Florida Business Owners Networking Group",
     "https://www.facebook.com/groups/sflbusinessowners/",
     "Miami-Dade + Broward",
     "South Florida business owners",
     "Marketplace: list any commercial space",
     "SoFL networking-focused.",
     HOST_URL, IMG_HOST_MIAMI,
     "Hey SoFL owners, my wife and I just launched alignworkspaces.com/for-hosts and wanted to share it with this group.\n\n"
     "If you've got commercial space in Miami or Broward that isn't in use 24/7, you can list those empty hours and earn from them. Free to list, no monthly.\n\n"
     "Open to feedback from anyone in this network. Mods, take this down if it isn't allowed."),

    ("Coral Springs Business Group",
     "https://www.facebook.com/groups/CoralSpringsBusinessGroup/",
     "Broward (Coral Springs)",
     "Coral Springs business owners",
     "Marketplace: list any commercial space",
     "Hyperlocal Coral Springs.",
     HOST_URL, IMG_HOST_BROWARD,
     "Hey Coral Springs, my wife and I just launched alignworkspaces.com/for-hosts and wanted to share it here.\n\n"
     "If you own commercial space (office, treatment room, studio, meeting room) and it sits empty for parts of the week, you can list those hours and earn from them. No upfront cost.\n\n"
     "Wanted to put it in front of locals first. Mods, please remove if this isn't allowed."),

    ("Coral Springs / Coconut Creek / Tamarac / Margate / Sunrise / Davie",
     "https://www.facebook.com/groups/coralspringscoconutcreeksunrise/",
     "Broward (multi-city)",
     "Broward multi-city locals + business owners",
     "Marketplace: list any commercial space",
     "Covers half of central Broward.",
     HOST_URL, IMG_HOST_BROWARD,
     "Hey Broward, my wife and I just launched alignworkspaces.com/for-hosts and wanted to share it with this group.\n\n"
     "If you own commercial space anywhere in the Coral Springs / Coconut Creek / Tamarac / Margate / Sunrise / Davie area and it sits unused part of the week, you can list those hours and earn from them. Free to list, no monthly.\n\n"
     "Wanted to put it in front of locals first. Mods, take this down if it isn't allowed."),

    ("Coral Gables Networking (Rockstar Connect)",
     "https://www.facebook.com/groups/1530498090397196/",
     "Miami-Dade (Coral Gables)",
     "Coral Gables business networking",
     "Marketplace: list any commercial space",
     "Networking-focused Coral Gables group.",
     HOST_URL, IMG_HOST_MIAMI,
     "Hey Coral Gables, my wife and I just launched alignworkspaces.com/for-hosts and wanted to share it with this group.\n\n"
     "If you own commercial space in Coral Gables (or anywhere in Miami) that sits empty part of the week, you can list those hours and earn from them. No upfront cost.\n\n"
     "Wanted to put it in front of this network first. Mods, please remove if it isn't allowed."),

    ("CG Business Connections (Coral Gables)",
     "https://www.facebook.com/groups/565526383779835/",
     "Miami-Dade (Coral Gables)",
     "Coral Gables business connections",
     "Marketplace: list any commercial space",
     "Active Coral Gables business group.",
     HOST_URL, IMG_HOST_MIAMI,
     "Hey CG, my wife and I just launched alignworkspaces.com/for-hosts and wanted to share it here.\n\n"
     "If you've got commercial space in the Gables or anywhere in Miami and it sits empty part of the week, you can list those hours and earn from them. Free to list, no monthly.\n\n"
     "Wanted to put it in front of this group for feedback. Mods, take this down if it isn't allowed."),

    ("Aventura Florida - City of Excellence",
     "https://www.facebook.com/groups/AventuraCityofExcellence/",
     "Miami-Dade (Aventura)",
     "Aventura locals + business community",
     "Marketplace: list any commercial space",
     "Community group, includes many Aventura biz owners.",
     HOST_URL, IMG_HOST_MIAMI,
     "Hey Aventura, my wife and I just launched alignworkspaces.com/for-hosts and wanted to share it with the community.\n\n"
     "If you own commercial space in Aventura and it sits empty for parts of the week, you can list those hours and earn from them. No upfront cost, no monthly.\n\n"
     "Wanted to put it in front of locals first. Mods, please remove if this isn't allowed."),
]

# === Sheet 2: NATIONAL groups for SaaS Studio Software pitch ===
SAAS_GROUPS = [
    # --- Salon / Spa / Beauty ---
    ("Salon Owners Support (S.O.S.)",
     "https://www.facebook.com/groups/1580311542183792/",
     "Salon / Beauty", "Salon owners (national)",
     "SaaS: replace stitched-together booking tools",
     "One of the larger salon owner communities.",
     SAAS_URL, IMG_SAAS_SALON,
     "Hey salon owners, my wife and I just launched alignworkspaces.com/for-studios and wanted to share it with this group.\n\n"
     "It's booking software built for salon owners who already have their own clients. Branded booking page under your own domain, calendar with double-booking prevention, Stripe payouts direct to you, automated reminders. No commission on bookings, you keep 100%.\n\n"
     "We built it because we kept hearing from salon owner friends that they were stitching together Square, Calendly, and email reminders, and it never quite worked. Open to any feedback. Mods, please remove if this isn't allowed."),

    ("Salon Owners United",
     "https://www.facebook.com/groups/499682997391285/",
     "Salon / Beauty", "Salon owners (national)",
     "SaaS: branded booking, no per-booking commission",
     "Active community.",
     SAAS_URL, IMG_SAAS_SALON,
     "Hey everyone, my wife and I just launched alignworkspaces.com/for-studios and wanted to put it in front of this community.\n\n"
     "It's booking and scheduling software for salon owners with their own clients. Your own branded booking page, calendar, contracts, automated arrival reminders, direct Stripe payouts. 14-day free trial, $29/mo, no per-booking commission.\n\n"
     "Built it after watching too many talented salon owners glue together five different tools to run one business. Would love feedback from people who actually do this for a living. Mods, take this down if not allowed."),

    ("Salon Business Owners Only",
     "https://www.facebook.com/groups/salonownersonly/",
     "Salon / Beauty", "Salon owners (national)",
     "SaaS: 0% commission, branded booking",
     "Strict owners-only filter.",
     SAAS_URL, IMG_SAAS_SALON,
     "Hey owners, my wife and I just launched alignworkspaces.com/for-studios and wanted to share with this group.\n\n"
     "Quick rundown: branded booking page (your logo, your colors), calendar that prevents double-booking, direct Stripe payouts (we never touch your money), contracts, arrival guides, automated reminders. Flat $29 to $299 a month, 0% commission.\n\n"
     "Built it because the existing options either take a cut of every booking or feel like they were designed for hair salons in 2010. Open to feedback. Mods, please remove if this doesn't fit."),

    ("Salon Suites and Booths",
     "https://www.facebook.com/groups/salonsuitesandbooths/",
     "Salon / Beauty", "Salon suite + booth renters/owners (national)",
     "SaaS: software for suite owners with own clients",
     "Highly relevant, suite owners are exact ICP.",
     SAAS_URL, IMG_SAAS_SALON,
     "Hey suite owners, my wife and I just launched alignworkspaces.com/for-studios and wanted to share it here.\n\n"
     "If you've got your own clients and just need clean booking software (branded page, calendar, payments, reminders), this might fit. Flat monthly fee, 0% commission, you keep all your booking money.\n\n"
     "We built it specifically because suite owners told us they hated either paying booking platforms a cut, or paying chair-management software for features they don't use. Mods, take this down if not allowed."),

    ("Profitable Salon & Spa Owners",
     "https://www.facebook.com/groups/profitablesalonowners/",
     "Salon / Beauty", "Salon + spa owners (national)",
     "SaaS: $29-$299/mo flat, 0% commission",
     "Profit-focused community.",
     SAAS_URL, IMG_SAAS_SALON,
     "Hey owners, my wife and I just launched alignworkspaces.com/for-studios and wanted to share with this group.\n\n"
     "Booking software for salon and spa owners with their own clients. Branded booking page, calendar, Stripe payouts direct to you, automated reminders, contracts. $29 to $299/mo flat, 0% commission per booking.\n\n"
     "Built it because most existing options either took a cut of every booking or were way too complicated for what we needed. Open to feedback from anyone running a profitable shop. Mods, please remove if this isn't allowed."),

    ("Salon and Spa Business Owners",
     "https://www.facebook.com/groups/302877396417262/",
     "Salon / Beauty", "Salon + spa business owners (national)",
     "SaaS: branded booking, no commission",
     "Active group.",
     SAAS_URL, IMG_SAAS_SALON,
     "Hey everyone, my wife and I just launched alignworkspaces.com/for-studios and wanted to put it in front of this group.\n\n"
     "Booking software for salon and spa owners with their own client base. Your branded page, calendar, payments, reminders. Flat monthly fee, no commission per booking, 14-day free trial.\n\n"
     "Open to feedback from anyone who runs a salon or spa. Mods, take this down if it doesn't fit."),

    ("Salon Owners Group",
     "https://www.facebook.com/groups/371059684270768/",
     "Salon / Beauty", "Salon owners (smaller national group)",
     "SaaS: branded booking, 0% commission",
     "Smaller engaged group.",
     SAAS_URL, IMG_SAAS_SALON,
     "Hey owners, my wife and I just launched alignworkspaces.com/for-studios and wanted to share it here.\n\n"
     "It's booking software for salon owners who already have their own clients. Branded booking page, calendar with double-booking prevention, Stripe payouts straight to you, automated client reminders. Flat $29 to $299/mo, no per-booking commission.\n\n"
     "Built after watching too many salon owners glue together Square + Calendly + email reminders. Open to feedback. Mods, please remove if this isn't allowed."),

    ("Salon Owners Support Group",
     "https://www.facebook.com/groups/salonownerssupportgroup/",
     "Salon / Beauty", "Salon owners (support-focused, national)",
     "SaaS: cleaner stack than current options",
     "Support-focused.",
     SAAS_URL, IMG_SAAS_SALON,
     "Hey everyone, my wife and I just launched alignworkspaces.com/for-studios and wanted to put it in front of this group.\n\n"
     "Booking + scheduling software for salon owners with their own clients. Branded booking page, calendar, Stripe payouts, contracts, reminders. Flat monthly fee, 0% commission per booking.\n\n"
     "Open to honest feedback from anyone running a salon. Mods, take this down if it isn't allowed."),

    ("Salon Owners Exclusive Networking Forum",
     "https://www.facebook.com/groups/138466176314119/",
     "Salon / Beauty", "Salon owners (smaller, exclusive)",
     "SaaS: $29-$299/mo flat, 0% commission",
     "Smaller exclusive group.",
     SAAS_URL, IMG_SAAS_SALON,
     "Hey owners, my wife and I just launched alignworkspaces.com/for-studios and wanted to share with this group.\n\n"
     "Booking software built for salon owners with their own client base. Your own branded page, calendar that prevents double-booking, direct Stripe payouts, contracts, automated reminders. Flat $29 to $299/mo, no per-booking commission.\n\n"
     "Wanted to put it in front of this community for feedback. Mods, please remove if this isn't allowed."),

    ("Booth Rental Listings",
     "https://www.facebook.com/groups/825524130880289/",
     "Salon / Beauty", "Salon booth + suite renters (national)",
     "SaaS: software for booth renters with own clients",
     "Booth/suite renters paying booking platforms on top of booth rent.",
     SAAS_URL, IMG_SAAS_SALON,
     "Hey booth renters, my wife and I just launched alignworkspaces.com/for-studios and wanted to share it here.\n\n"
     "If you're renting a booth and have your own client base, this gives you a branded booking page, calendar, Stripe payouts, contracts, reminders. $29/mo flat, 0% commission, you keep all your booking money.\n\n"
     "Built specifically because booth renters told us they hated paying booking platforms a percentage on top of their booth rent. Mods, please remove if this isn't allowed."),

    # --- Med Spa / Aesthetics ---
    ("Medical Spa Owners",
     "https://www.facebook.com/groups/medicalspaowner/",
     "Med Spa / Aesthetics", "Med spa owners (national)",
     "SaaS: branded booking, intake, no commission",
     "Med spa owners often paying $500+/mo.",
     SAAS_URL, IMG_SAAS_MEDSPA,
     "Hey med spa owners, my wife and I just launched alignworkspaces.com/for-studios and wanted to share with this group.\n\n"
     "It's booking software for service practices with their own clients. Branded booking page under your domain, calendar, intake forms, contracts, direct Stripe payouts, automated reminders. $29 to $299/mo flat, no booking commission.\n\n"
     "Built it after hearing too many med spa owners say their booking system was either expensive ($500+/mo) or felt like it was made for hair salons. Would love feedback. Mods, please remove if this isn't allowed."),

    ("MedSpa & Clinic Owners Group",
     "https://www.facebook.com/groups/medspaownersgroup/",
     "Med Spa / Aesthetics", "Med spa + clinic owners (national)",
     "SaaS: simpler than legacy clinic software",
     "Engaged owner community.",
     SAAS_URL, IMG_SAAS_MEDSPA,
     "Hey clinic owners, my wife and I just launched alignworkspaces.com/for-studios and wanted to put it in front of this group.\n\n"
     "Booking software built for service practices: branded booking page, calendar, intake forms, contracts, Stripe payouts direct to you. Flat monthly fee, 0% commission, 14-day free trial.\n\n"
     "Open to feedback if anyone wants to take a look. Mods, please remove if this isn't a good fit."),

    ("Med Spa Owners Community",
     "https://www.facebook.com/groups/medspaownerscommunity/",
     "Med Spa / Aesthetics", "Med spa owners (national)",
     "SaaS: replace fragmented booking stack",
     "Active.",
     SAAS_URL, IMG_SAAS_MEDSPA,
     "Hey everyone, my wife and I just launched alignworkspaces.com/for-studios and wanted to share it with this community.\n\n"
     "It's a booking + scheduling system for clinics and service practices that already have their own clients. Your branded page, calendar, payments, intake. Flat monthly, 0% per-booking fee.\n\n"
     "Built because we watched too many med spa owners stitch together five tools for one workflow. Would love your feedback. Mods, take this down if not allowed."),

    ("Cosmetic Laser Clinic and Med Spa Owners Group",
     "https://www.facebook.com/groups/123632434893396/",
     "Med Spa / Aesthetics", "Cosmetic laser + med spa owners (national)",
     "SaaS: cheaper than $500+/mo legacy clinic software",
     "Niche-specific aesthetics group.",
     SAAS_URL, IMG_SAAS_MEDSPA,
     "Hey clinic owners, my wife and I just launched alignworkspaces.com/for-studios and wanted to put it in front of this group.\n\n"
     "Booking + intake software for med spas and laser clinics. Branded booking page, calendar, intake forms, contracts, direct Stripe payouts, automated reminders. Flat monthly, 0% per-booking commission.\n\n"
     "Built it because the existing options were either expensive ($500+/mo) or felt like they were designed for hair salons. Open to feedback. Mods, take this down if not allowed."),

    ("Medical Spa Connect",
     "https://www.facebook.com/groups/351899701903135/",
     "Med Spa / Aesthetics", "Med spa owners (national)",
     "SaaS: branded booking, 0% commission",
     "Active med spa community.",
     SAAS_URL, IMG_SAAS_MEDSPA,
     "Hey everyone, my wife and I just launched alignworkspaces.com/for-studios and wanted to share with this community.\n\n"
     "Booking and payment software for med spas with their own clients. Branded booking page, calendar, intake forms, Stripe payouts. Flat monthly fee, 0% per-booking commission.\n\n"
     "Open to feedback from anyone running a med spa. Mods, please remove if this isn't allowed."),

    # --- Fitness / Studios ---
    ("Pilates & Yoga Studio Owners Network",
     "https://www.facebook.com/groups/studioownersnetwork/",
     "Fitness / Studios", "Pilates + yoga studio owners (national)",
     "SaaS: cheaper, simpler than Mindbody",
     "Direct ICP, studio owners hate Mindbody pricing.",
     SAAS_URL, IMG_SAAS_FITNESS,
     "Hey studio owners, my wife and I just launched alignworkspaces.com/for-studios and wanted to share it with this group.\n\n"
     "It's booking + scheduling software for studios with their own clients. Branded booking page, calendar with double-booking prevention, direct Stripe payouts, automated class reminders. $29 to $299/mo flat, 0% per-booking commission.\n\n"
     "Built after watching studio owner friends glue together Mindbody, Square, and Mailchimp for one workflow. Would love feedback from anyone who actually runs a studio. Mods, please remove if this isn't allowed."),

    ("Pilates Studio Owners Collaboration for Profit",
     "https://www.facebook.com/groups/2019805464943774/",
     "Fitness / Studios", "Pilates studio owners (national)",
     "SaaS: flat fee, 0% commission",
     "Profit-focused community.",
     SAAS_URL, IMG_SAAS_FITNESS,
     "Hey pilates studio owners, my wife and I just launched alignworkspaces.com/for-studios and wanted to put it in front of this group.\n\n"
     "Booking software built for studio operators who already have a client base. Branded booking page, calendar, payments, contracts, reminders. Flat monthly fee, no commission per class.\n\n"
     "Built because everyone we know running a studio said the existing options were either too expensive or way too complicated. Open to any feedback. Mods, take this down if it doesn't fit."),

    ("Gym Owner Support Group",
     "https://www.facebook.com/groups/fitproleadgen/",
     "Fitness / Studios", "Gym + fitness business owners (national)",
     "SaaS: booking + payments without per-booking cut",
     "Gym owners often using bloated all-in-one software.",
     SAAS_URL, IMG_SAAS_FITNESS,
     "Hey gym owners, my wife and I just launched alignworkspaces.com/for-studios and wanted to share with this group.\n\n"
     "It's booking + payment software for fitness businesses with their own clients. Branded booking page, schedule, payments via Stripe, automated reminders. Flat $29 to $299/mo, 0% per-booking commission.\n\n"
     "Built it after hearing the same complaint over and over: 'my booking software costs more than my rent.' Would love feedback. Mods, please remove if this isn't allowed."),

    ("Personal Trainers & Fitness Instructors",
     "https://www.facebook.com/groups/personaltrainersandfitnessinstructors/",
     "Fitness / Studios", "Personal trainers + fitness pros (national)",
     "SaaS: branded booking for trainers with own clients",
     "Solo trainers often running on DMs + Venmo.",
     SAAS_URL, IMG_SAAS_FITNESS,
     "Hey trainers, my wife and I just launched alignworkspaces.com/for-studios and wanted to share it here.\n\n"
     "Booking software for trainers and fitness pros with their own clients. Branded page (your logo, your colors), calendar, Stripe payouts, automated client reminders. Flat monthly fee, 0% commission on bookings.\n\n"
     "Built because most trainer friends were either paying ClassPass-style platforms a cut or running everything through DMs and Venmo. Open to your honest feedback. Mods, take this down if not allowed."),

    ("Personal Trainers & Fitness Professional Support Group",
     "https://www.facebook.com/groups/personaltrainersandfitnessprofessionalsupportgroup/",
     "Fitness / Studios", "Personal trainers + fitness pros (national)",
     "SaaS: branded booking + payments for trainers",
     "Larger trainer community.",
     SAAS_URL, IMG_SAAS_FITNESS,
     "Hey trainers, my wife and I just launched alignworkspaces.com/for-studios and wanted to share with this group.\n\n"
     "Booking + payment software for fitness pros with their own clients. Your branded booking page, calendar, Stripe payouts direct to you, automated reminders. Flat $29 to $299/mo, 0% commission per session.\n\n"
     "Built it because most trainer friends were either using free Calendly + Venmo (no follow-through) or paying ClassPass-style platforms a cut. Open to feedback. Mods, take this down if not allowed."),

    # --- Dance ---
    ("Dance Studio OWNERS",
     "https://www.facebook.com/groups/dancestudioowners/",
     "Dance", "Dance studio owners (national)",
     "SaaS: simpler than legacy dance studio software",
     "Active dance studio owner community.",
     SAAS_URL, IMG_SAAS_FITNESS,
     "Hey studio owners, my wife and I just launched alignworkspaces.com/for-studios and wanted to share it with this group.\n\n"
     "Booking + payment software for dance studios with their own clients. Branded booking page, schedule, payments, automated reminders. $29 to $299/mo flat, no per-booking commission.\n\n"
     "Built it after hearing studio owner friends complain that their existing platforms either took a percentage or felt overdesigned for what they actually needed. Would love feedback from anyone who runs a studio. Mods, please remove if this isn't a fit."),

    ("Dance Studio Owner's Association [DSOA]",
     "https://www.facebook.com/groups/thedsoa/",
     "Dance", "DSOA member studio owners (national)",
     "SaaS: flat fee booking software",
     "Affiliated with the DSOA membership org.",
     SAAS_URL, IMG_SAAS_FITNESS,
     "Hey DSOA folks, my wife and I just launched alignworkspaces.com/for-studios and wanted to put it in front of this group.\n\n"
     "It's booking and scheduling software for studios with their own client base. Branded page, calendar, Stripe payouts, contracts, reminders, all flat monthly fee with 0% per-booking commission.\n\n"
     "Open to feedback from anyone here who's tested booking platforms. Mods, take this down if it isn't allowed."),

    # --- Therapy / Mental Health ---
    ("Therapists in Private Practice (TIPP)",
     "https://www.facebook.com/groups/TIPPgroup/",
     "Therapy / Mental Health", "Therapists in private practice (national)",
     "SaaS: booking that doesn't take a cut of sessions",
     "Large private-practice therapist community.",
     SAAS_URL, IMG_SAAS_THERAPY,
     "Hey therapists, my wife and I just launched alignworkspaces.com/for-studios and wanted to share with this group.\n\n"
     "It's booking software for private practice that doesn't take a cut of your sessions. Branded booking page, calendar, payments via Stripe direct to you, automated reminders, intake forms. Flat $29 to $299/mo.\n\n"
     "Built after watching too many therapist friends pay 5 to 10% of every session to their booking platform. Would love feedback from anyone in private practice. Mods, please remove if this isn't allowed."),

    ("Online Therapists Group",
     "https://www.facebook.com/groups/138663903332494/",
     "Therapy / Mental Health", "Online + hybrid therapists (national)",
     "SaaS: 0% on sessions, branded booking",
     "Online-first therapists.",
     SAAS_URL, IMG_SAAS_THERAPY,
     "Hey therapists, my wife and I just launched alignworkspaces.com/for-studios and wanted to share it with this group.\n\n"
     "Booking software for online and in-person therapy that doesn't take a cut of your sessions. Branded booking page, calendar, Stripe payouts direct to you, automated reminders, intake forms. Flat $29 to $299/mo.\n\n"
     "Built after watching therapist friends pay 5 to 10% of every session to their booking platform. Open to feedback. Mods, please remove if not allowed."),

    # --- Coaches ---
    ("Extraordinary Coaches & Business Owners",
     "https://www.facebook.com/groups/extraordinarycoaches/",
     "Coaching", "Coaches + service business owners (national)",
     "SaaS: booking + Stripe payouts for coaches",
     "Coach + business owner audience.",
     SAAS_URL, IMG_SAAS_COACH,
     "Hey coaches, my wife and I just launched alignworkspaces.com/for-studios and wanted to share it with this group.\n\n"
     "Booking software for coaches with their own clients. Branded booking page (your logo, your colors), calendar, Stripe payouts straight to you, automated reminders, contracts. $29 to $299/mo flat, 0% per-session commission.\n\n"
     "Built after watching coach friends try to MacGyver Calendly + Stripe + email reminders. Open to your honest feedback. Mods, please remove if this isn't allowed."),

    ("High Ticket Coaches - Business Owners - Entrepreneurs",
     "https://www.facebook.com/groups/1451795975240926/",
     "Coaching", "High-ticket coaches + consultants (national)",
     "SaaS: 0% on high-ticket sessions",
     "Especially compelling for high-ticket where % cuts hurt most.",
     SAAS_URL, IMG_SAAS_COACH,
     "Hey everyone, my wife and I just launched alignworkspaces.com/for-studios and wanted to put it in front of this group.\n\n"
     "Booking + payment software for coaches and consultants with their own client base. Branded booking page, calendar, contracts, direct Stripe payouts. Flat monthly fee, 0% per-booking commission. No taking a cut of your high-ticket sessions.\n\n"
     "Open to feedback from anyone who's tested booking platforms. Mods, take this down if not allowed."),

    ("Life Coaching Community",
     "https://www.facebook.com/groups/1097628144155616/",
     "Coaching", "Life coaches (national)",
     "SaaS: replace Calendly + Stripe + email stack",
     "Active coaching community.",
     SAAS_URL, IMG_SAAS_COACH,
     "Hey coaches, my wife and I just launched alignworkspaces.com/for-studios and wanted to share with this community.\n\n"
     "Booking and scheduling software built for coaches and service practitioners. Branded page, calendar, payments, reminders. Flat monthly fee, no per-session commission, 14-day free trial.\n\n"
     "Built it because too many coach friends were trying to glue together Calendly + Stripe + email + spreadsheets. Would love feedback. Mods, please remove if this isn't allowed."),

    ("LIFE COACH",
     "https://www.facebook.com/groups/1669765473317555/",
     "Coaching", "Life coaches (national)",
     "SaaS: replace Calendly + Stripe + email stack",
     "Active life coach community.",
     SAAS_URL, IMG_SAAS_COACH,
     "Hey coaches, my wife and I just launched alignworkspaces.com/for-studios and wanted to put it in front of this group.\n\n"
     "Booking + scheduling software for coaches with their own clients. Your branded booking page, calendar, Stripe payouts straight to you, contracts, automated reminders. Flat monthly fee, 0% per-session commission.\n\n"
     "Built it because too many coach friends were trying to glue together Calendly + Stripe + email + spreadsheets. Open to feedback. Mods, take this down if not allowed."),

    # --- Massage / Bodywork ---
    ("Massage Therapists",
     "https://www.facebook.com/groups/611431842200765/",
     "Massage / Bodywork", "Massage therapists (national)",
     "SaaS: 0% per-booking, branded booking page",
     "Most massage software takes a cut per booking.",
     SAAS_URL, IMG_SAAS_THERAPY,
     "Hey massage therapists, my wife and I just launched alignworkspaces.com/for-studios and wanted to share it with this group.\n\n"
     "Booking software for massage practices with their own clients. Branded booking page, calendar with double-booking prevention, direct Stripe payouts, automated client reminders, intake. Flat $29/mo, 0% per-booking commission.\n\n"
     "Built it after watching too many bodyworkers get charged 5 to 10% per booking by their software. Open to feedback. Mods, please remove if this isn't allowed."),

    # --- Chiropractic ---
    ("The Smart Chiropractor",
     "https://www.facebook.com/groups/1815001008826046/",
     "Chiropractic / Wellness", "Chiropractors (national)",
     "SaaS: clinic booking without legacy software cost",
     "Most chiro software is bloated and expensive.",
     SAAS_URL, IMG_SAAS_MEDSPA,
     "Hey chiropractors, my wife and I just launched alignworkspaces.com/for-studios and wanted to put it in front of this group.\n\n"
     "Booking + intake software for clinics with their own clients. Branded booking page, calendar, intake forms, Stripe payouts direct to you, automated reminders. Flat monthly, 0% per-booking commission.\n\n"
     "Built it because most chiro booking software either costs $500+/mo or feels like it was designed in 2008. Would love feedback. Mods, take this down if not allowed."),

    # --- Recording / Audio ---
    ("The Home Recording Studio Zone",
     "https://www.facebook.com/groups/audioproduction101/",
     "Recording / Audio", "Recording + podcast studio owners (national)",
     "SaaS: branded booking for studio operators",
     "Studio operators stitching together Calendly + Stripe.",
     SAAS_URL, IMG_SAAS_RECORDING,
     "Hey studio owners, my wife and I just launched alignworkspaces.com/for-studios and wanted to share it with this group.\n\n"
     "If you run a recording or podcast studio and have your own clients, this gives you a branded booking page, calendar, payments via Stripe, automated reminders. Flat $29 to $299/mo, no per-session commission.\n\n"
     "Built it because we watched studio operators stitch together Calendly, Stripe, and email reminders to run one workflow. Open to your feedback. Mods, please remove if this isn't allowed."),

    # --- Acupuncture / Wellness ---
    ("Acupuncturists on Facebook",
     "https://www.facebook.com/groups/22073416402/",
     "Acupuncture / Wellness", "Acupuncturists (national)",
     "SaaS: branded booking + intake for solo practitioners",
     "Specific acupuncture practitioner community.",
     SAAS_URL, IMG_SAAS_THERAPY,
     "Hey acupuncturists, my wife and I just launched alignworkspaces.com/for-studios and wanted to share it with this group.\n\n"
     "Booking software for acupuncture practices with their own clients. Branded booking page (your domain), calendar, Stripe payouts, intake forms, automated reminders. Flat $29 to $299/mo, 0% per-session commission.\n\n"
     "Built it because most practice software either takes a cut of sessions or feels designed for medical clinics, not solo practitioners. Open to feedback. Mods, take this down if not allowed."),

    # --- Wedding / Photography ---
    ("Wedding Photography & Videography",
     "https://www.facebook.com/groups/weddingphotographyandvideography/",
     "Wedding / Photography", "Wedding photographers + videographers (national)",
     "SaaS: branded booking + contracts for wedding pros",
     "Wedding pros stitching HoneyBook + Calendly + Stripe.",
     SAAS_URL, IMG_SAAS_PHOTO,
     "Hey wedding pros, my wife and I just launched alignworkspaces.com/for-studios and wanted to share with this group.\n\n"
     "Booking + payment software for wedding photographers and videographers with their own clients. Branded booking page, calendar, contracts, Stripe payouts direct to you, automated client reminders. Flat $29 to $299/mo, 0% commission per booking.\n\n"
     "Built it because too many wedding pros were stitching together HoneyBook + Calendly + Stripe + email reminders for one workflow. Open to feedback. Mods, please remove if this isn't allowed."),

    # --- Nutrition / Dietitians ---
    ("Dietitians in Private Practice",
     "https://www.facebook.com/groups/927575603995649/",
     "Nutrition / Dietitians", "Private practice dietitians (national)",
     "SaaS: branded booking + intake, no per-client fees",
     "Specific to private practice dietitians.",
     SAAS_URL, IMG_SAAS_COACH,
     "Hey dietitians, my wife and I just launched alignworkspaces.com/for-studios and wanted to put it in front of this group.\n\n"
     "Booking software for private practice dietitians. Branded booking page, calendar, Stripe payouts, intake forms, automated reminders. Flat $29 to $299/mo, 0% per-booking commission.\n\n"
     "Built it because most practice software either takes a cut of sessions or charges per-client fees. Open to feedback from anyone in private practice. Mods, take this down if not allowed."),

    ("Dietitians in Group Practice",
     "https://www.facebook.com/groups/dietitiansingrouppractice/",
     "Nutrition / Dietitians", "Group practice dietitians (national)",
     "SaaS: multi-practitioner calendar + branded booking",
     "Group practice angle.",
     SAAS_URL, IMG_SAAS_COACH,
     "Hey everyone, my wife and I just launched alignworkspaces.com/for-studios and wanted to share it here.\n\n"
     "Booking + scheduling software for group nutrition practices. Branded booking page, multi-practitioner calendar, Stripe payouts, intake forms. Flat monthly fee, no per-booking commission.\n\n"
     "Open to feedback from anyone running a group practice. Mods, please remove if this isn't allowed."),

    # --- Music / Education ---
    ("International Music Teachers EXchange",
     "https://www.facebook.com/groups/362815390754879/",
     "Music / Education", "Music teachers (national / global)",
     "SaaS: branded booking + payments for private teachers",
     "Music teachers often using DM + Venmo workflow.",
     SAAS_URL, IMG_SAAS_RECORDING,
     "Hey teachers, my wife and I just launched alignworkspaces.com/for-studios and wanted to share with this group.\n\n"
     "Booking + payment software for music teachers with their own students. Branded booking page, calendar, Stripe payouts direct to you, automated reminders. Flat $29 to $299/mo, no per-booking commission.\n\n"
     "Built after watching teacher friends try to MacGyver Calendly + Venmo + email reminders for one workflow. Open to feedback. Mods, take this down if not allowed."),

    ("Music Teachers (community)",
     "https://www.facebook.com/groups/musicpln/",
     "Music / Education", "Music teachers (national)",
     "SaaS: branded booking + reminders for private teachers",
     "Active music teacher community.",
     SAAS_URL, IMG_SAAS_RECORDING,
     "Hey everyone, my wife and I just launched alignworkspaces.com/for-studios and wanted to put it in front of this group.\n\n"
     "Booking + scheduling software for private music teachers. Your branded booking page, calendar, Stripe payouts, automated student reminders, contracts. Flat monthly fee, 0% per-lesson commission.\n\n"
     "Open to feedback from anyone teaching privately. Mods, please remove if this isn't allowed."),

    # --- Holistic / Wellness ---
    ("Holistic Health Practitioners",
     "https://www.facebook.com/groups/holistichealthpractitionersoGC/",
     "Holistic / Wellness", "Holistic health practitioners (national)",
     "SaaS: branded booking + intake for solo practitioners",
     "Solo holistic practitioners.",
     SAAS_URL, IMG_SAAS_THERAPY,
     "Hey practitioners, my wife and I just launched alignworkspaces.com/for-studios and wanted to share it with this group.\n\n"
     "Booking software for holistic practitioners with their own clients. Branded booking page (your domain, your colors), calendar, Stripe payouts, intake forms, automated reminders. Flat $29 to $299/mo, 0% per-session commission.\n\n"
     "Built it because most practitioner software either takes a cut of sessions or is designed for medical clinics rather than holistic practice. Open to feedback. Mods, take this down if not allowed."),

    # === New additions (round 3) ===
    # --- Tattoo ---
    ("Tattoo Apprentices and Artists",
     "https://www.facebook.com/groups/415455009041433/",
     "Tattoo / Body Art", "Tattoo artists + shop owners (national)",
     "SaaS: branded booking + deposits for tattoo artists",
     "Tattoo artists often using DMs + Venmo for bookings.",
     SAAS_URL, IMG_SAAS_SALON,
     "Hey artists, my wife and I just launched alignworkspaces.com/for-studios and wanted to share it with this group.\n\n"
     "If you run your own tattoo book and have your own clients, this gives you a branded booking page, calendar, deposits via Stripe, intake forms, automated reminders. Flat $29 to $299/mo, 0% commission per booking.\n\n"
     "Built because too many artists we know were running their entire booking through DMs + Venmo and getting flaked on. Open to feedback. Mods, please remove if not allowed."),

    # --- Lash artists ---
    ("Lash Tech Help Group / Community",
     "https://www.facebook.com/groups/937034773648940/",
     "Lash / Brow", "Lash artists + lash business owners (national)",
     "SaaS: branded booking + reminders for lash artists",
     "Active lash artist community.",
     SAAS_URL, IMG_SAAS_SALON,
     "Hey lash artists, my wife and I just launched alignworkspaces.com/for-studios and wanted to share it with this group.\n\n"
     "Booking software for lash artists with their own clients. Branded booking page, calendar, Stripe payouts direct to you, automated reminders, intake forms. Flat $29/mo, 0% per-booking commission.\n\n"
     "Built it because most existing options either take a cut of every booking or are way too overkill for what a lash business actually needs. Open to feedback. Mods, take this down if not allowed."),

    ("Lash Community (Eyelash Extension Artists)",
     "https://www.facebook.com/groups/eyelashextensionartistsgroup/",
     "Lash / Brow", "Eyelash extension artists (national)",
     "SaaS: branded booking + intake for lash pros",
     "Specific to extension artists.",
     SAAS_URL, IMG_SAAS_SALON,
     "Hey everyone, my wife and I just launched alignworkspaces.com/for-studios and wanted to put it in front of this group.\n\n"
     "Booking software built for lash artists with their own client base. Branded page (your logo, your colors), calendar, Stripe payouts, intake, reminders. Flat monthly fee, no per-booking commission.\n\n"
     "Open to feedback from anyone running a lash business. Mods, please remove if this isn't allowed."),

    # --- Barbershop ---
    ("Barbershop Chat",
     "https://www.facebook.com/groups/barbershopchat/",
     "Barbershop", "Barbershop owners + barbers (national)",
     "SaaS: branded booking + reminders for barbershops",
     "Active barbershop community.",
     SAAS_URL, IMG_SAAS_SALON,
     "Hey barbers, my wife and I just launched alignworkspaces.com/for-studios and wanted to share it with this group.\n\n"
     "Booking software for barbershops + chair-renting barbers with their own clients. Branded booking page, calendar, Stripe payouts direct to you, automated reminders. Flat $29 to $299/mo, 0% per-booking commission.\n\n"
     "Built it because most barbershop software either takes a cut of every booking or feels overdesigned for what a shop actually needs. Open to feedback. Mods, please remove if this isn't allowed."),

    ("Professional Barbers",
     "https://www.facebook.com/groups/ProfessionalBarbers/",
     "Barbershop", "Professional barbers (national)",
     "SaaS: branded booking, no per-cut commission",
     "Professional barber community.",
     SAAS_URL, IMG_SAAS_SALON,
     "Hey everyone, my wife and I just launched alignworkspaces.com/for-studios and wanted to put it in front of this group.\n\n"
     "Booking + payment software for barbers with their own client book. Branded booking page, calendar, Stripe payouts, automated reminders. Flat monthly fee, no commission per cut.\n\n"
     "Open to feedback from anyone behind the chair. Mods, take this down if not allowed."),

    # --- Event Planners / Wedding Planners ---
    ("Event Planners Society",
     "https://www.facebook.com/groups/eventplannerssociety/",
     "Event / Wedding Planning", "Event planners (national)",
     "SaaS: branded booking + contracts for planners",
     "Event planning society.",
     SAAS_URL, IMG_SAAS_PHOTO,
     "Hey planners, my wife and I just launched alignworkspaces.com/for-studios and wanted to share it with this group.\n\n"
     "Booking + payment software for event planners with their own clients. Branded booking page, calendar, contracts, Stripe payouts direct to you, automated reminders. Flat $29 to $299/mo, 0% commission per booking.\n\n"
     "Built it because most planner friends were stitching together HoneyBook + Calendly + Stripe + email for one workflow. Open to feedback. Mods, please remove if this isn't allowed."),

    ("Weddings & Events Planner Group",
     "https://www.facebook.com/groups/127366957958976/",
     "Event / Wedding Planning", "Wedding + event planners (national)",
     "SaaS: branded booking + contracts for planners",
     "Wedding + event planner audience.",
     SAAS_URL, IMG_SAAS_PHOTO,
     "Hey planners, my wife and I just launched alignworkspaces.com/for-studios and wanted to put it in front of this group.\n\n"
     "Booking + scheduling software for wedding and event planners. Branded booking page, contracts, calendar, Stripe payouts, automated reminders. Flat monthly fee, 0% commission per booking.\n\n"
     "Open to feedback from anyone planning events for a living. Mods, take this down if not allowed."),

    # --- Sound healing / energy work ---
    ("Sound Bath Practitioners World Wide",
     "https://www.facebook.com/groups/235814466274117/",
     "Holistic / Wellness", "Sound bath + energy healing practitioners (national)",
     "SaaS: branded booking for solo practitioners",
     "Sound healers + energy workers with own clients.",
     SAAS_URL, IMG_SAAS_FITNESS,
     "Hey practitioners, my wife and I just launched alignworkspaces.com/for-studios and wanted to share it with this group.\n\n"
     "Booking software for solo practitioners with their own clients. Branded booking page (your domain, your colors), calendar, Stripe payouts, intake forms, automated reminders. Flat $29 to $299/mo, 0% per-session commission.\n\n"
     "Built it because most practice software either takes a cut of sessions or is designed for medical clinics rather than holistic + energy work. Open to feedback. Mods, take this down if not allowed."),

    # --- Tutoring ---
    ("Support for Private Tutoring Business Owners and Teachers",
     "https://www.facebook.com/groups/772680750374267/",
     "Tutoring / Education", "Tutoring business owners (national)",
     "SaaS: branded booking + payments for tutors",
     "Tutoring business owners.",
     SAAS_URL, IMG_SAAS_COACH,
     "Hey tutors, my wife and I just launched alignworkspaces.com/for-studios and wanted to share with this group.\n\n"
     "Booking + payment software for tutors with their own students. Branded booking page, calendar, Stripe payouts, automated reminders. Flat $29 to $299/mo, 0% per-booking commission.\n\n"
     "Built it because most tutor friends were running their entire business through DMs and Venmo + spreadsheets. Open to feedback. Mods, please remove if this isn't allowed."),

    # --- More music ---
    ("Private Music Instructors, Piano Lesson Teachers",
     "https://www.facebook.com/groups/privatemusicinstructors/",
     "Music / Education", "Private music instructors (national)",
     "SaaS: branded booking + payments for instructors",
     "Private music instructor community.",
     SAAS_URL, IMG_SAAS_RECORDING,
     "Hey instructors, my wife and I just launched alignworkspaces.com/for-studios and wanted to share it with this group.\n\n"
     "Booking + payment software for private music instructors. Branded booking page, calendar, Stripe payouts direct to you, automated student reminders, contracts. Flat $29 to $299/mo, 0% per-lesson commission.\n\n"
     "Built it because most instructor friends were trying to MacGyver Calendly + Venmo + email reminders. Open to feedback. Mods, take this down if not allowed."),

    ("Vibrant Music Studio Teachers",
     "https://www.facebook.com/groups/vibrantmusicstudioteachers/",
     "Music / Education", "Music studio teachers (national)",
     "SaaS: branded booking + reminders for music studios",
     "Music studio teacher community.",
     SAAS_URL, IMG_SAAS_RECORDING,
     "Hey teachers, my wife and I just launched alignworkspaces.com/for-studios and wanted to put it in front of this group.\n\n"
     "Booking + scheduling software for music studios with their own students. Your branded booking page, calendar, Stripe payouts, automated reminders, contracts. Flat monthly fee, 0% per-lesson commission.\n\n"
     "Open to feedback from anyone running a studio. Mods, please remove if this isn't allowed."),
]

# === Style helpers ===
HEADER_FILL = PatternFill("solid", fgColor="2A2A2A")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
LINK_FONT = Font(color="C4956A", underline="single", name="Calibri")
ALT_FILL = PatternFill("solid", fgColor="FAF8F5")
THIN = Side(border_style="thin", color="E5E0D8")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


POSTED_FILL = PatternFill("solid", fgColor="DCEEDC")  # soft green for posted rows
POSTED_CELL_FILL = PatternFill("solid", fgColor="6FB36F")  # stronger green in the checkbox cell
POSTED_FONT = Font(color="FFFFFF", bold=True, size=14, name="Calibri")


def write_sheet(ws, groups, category_label):
    headers = ["Posted", "#", "Group Name", "FB Group Link", category_label, "Audience",
               "Best Pitch Angle", "Notes", "Link to Share", "Image to Attach", "Post (copy/paste)"]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center" if col_idx == 1 else "left", vertical="center")
        cell.border = BORDER
    ws.row_dimensions[1].height = 28

    for i, row_data in enumerate(groups, start=1):
        name, url, cat, audience, pitch, notes, link_to_share, image, post = row_data
        row = i + 1
        ws.cell(row=row, column=1, value="")  # checkbox cell
        ws.cell(row=row, column=2, value=i)
        ws.cell(row=row, column=3, value=name)
        link_cell = ws.cell(row=row, column=4, value=url)
        link_cell.hyperlink = url
        link_cell.font = LINK_FONT
        ws.cell(row=row, column=5, value=cat)
        ws.cell(row=row, column=6, value=audience)
        ws.cell(row=row, column=7, value=pitch)
        ws.cell(row=row, column=8, value=notes)
        ws.cell(row=row, column=9, value=link_to_share)
        ws.cell(row=row, column=10, value=image)
        ws.cell(row=row, column=11, value=post)

        for col_idx in range(1, 12):
            c = ws.cell(row=row, column=col_idx)
            c.alignment = Alignment(wrap_text=True, vertical="top",
                                    horizontal="center" if col_idx == 1 else "left")
            c.border = BORDER
            if i % 2 == 0:
                c.fill = ALT_FILL

        # Make the checkbox cell big and centered
        check = ws.cell(row=row, column=1)
        check.alignment = Alignment(horizontal="center", vertical="center")
        check.font = POSTED_FONT

        ws.row_dimensions[row].height = 170

    # Dropdown so a single click marks "Posted" with a checkmark
    last_row = len(groups) + 1
    dv = DataValidation(type="list", formula1='"✓"', allow_blank=True,
                        showDropDown=False,
                        prompt="Click the dropdown arrow and pick ✓ to mark this group as posted.",
                        promptTitle="Mark as posted")
    dv.add(f"A2:A{last_row}")
    ws.add_data_validation(dv)

    # Conditional formatting: when col A has the checkmark, fill the whole row pale green
    # and the checkbox cell stronger green
    green_row_rule = FormulaRule(formula=[f'$A2="✓"'], stopIfTrue=False, fill=POSTED_FILL)
    ws.conditional_formatting.add(f"B2:K{last_row}", green_row_rule)

    green_cell_rule = FormulaRule(formula=[f'$A2="✓"'], stopIfTrue=False, fill=POSTED_CELL_FILL,
                                  font=POSTED_FONT)
    ws.conditional_formatting.add(f"A2:A{last_row}", green_cell_rule)

    widths = {1: 9, 2: 5, 3: 36, 4: 48, 5: 24, 6: 32, 7: 36, 8: 28, 9: 26, 10: 22, 11: 70}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "C2"


wb = Workbook()
ws1 = wb.active
ws1.title = "FB Groups (Hosts)"
write_sheet(ws1, HOST_GROUPS, "County")

ws2 = wb.create_sheet("SaaS Studio Software (National)")
write_sheet(ws2, SAAS_GROUPS, "Vertical")

# Save with fallback if Excel has the file open
import os
primary = "documents/marketing/fb-groups-to-post.xlsx"
fallback = "documents/marketing/fb-groups-to-post-v2.xlsx"
try:
    wb.save(primary)
    out = primary
except PermissionError:
    wb.save(fallback)
    out = fallback
    print(f"(Primary file was locked/open in Excel, wrote sibling: {fallback})")

print(f"Saved: {out}")
print(f"Sheet 1 (Hosts, Miami/Broward): {len(HOST_GROUPS)} groups")
print(f"Sheet 2 (SaaS, National): {len(SAAS_GROUPS)} groups")

from collections import Counter
print("\nSheet 1 by county:")
for v, c in sorted(Counter(g[2] for g in HOST_GROUPS).items(), key=lambda x: -x[1]):
    print(f"  {v}: {c}")
print("\nSheet 2 by vertical:")
for v, c in sorted(Counter(g[2] for g in SAAS_GROUPS).items(), key=lambda x: -x[1]):
    print(f"  {v}: {c}")
